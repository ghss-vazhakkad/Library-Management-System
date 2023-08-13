import sys
from PyQt5 import uic
from PyQt5.QtWidgets import *
from datetime import datetime

from openpyxl import Workbook, load_workbook



class BookEntry(QDialog):
    def __init__(self,parent):
        super(BookEntry, self).__init__()
        uic.loadUi('res/bookentry.ui', self)
        self.setFixedSize(self.size())
        self.parent = parent
        self.submit.clicked.connect(self.onadd)
        self.close_btn.clicked.connect(self.hide)
        self.show()
    def onadd(self):
        if self.dataTitle.text() != "" and self.dataID.text() != "":
            try:
                book = Book()
                book.date = [datetime.now().day,datetime.now().month,datetime.now().year]
                book.title = self.dataTitle.text()
                book.id = eval(self.dataID.text())
                book.author = self.dataAuthor.text()
                book.subject = self.dataSubject.text()
                book.language = self.dataLanguage.text()
                book.publisher = self.dataPublisher.text()
                if(self.dataYear.text() != ""): book.year = eval(self.dataYear.text())
                if(self.dataPrice.text() != ""): book.price = eval(self.dataPrice.text())
                if(self.radioReference.isChecked()):
                    book.booktype = Book.BOOKTYPE_REFERENCE
                else:
                    book.booktype = Book.BOOKTYPE_GENERAL
                if(self.radioGift.isChecked()):
                    book.etype = Book.TYPE_GIFT
                elif(self.radioDirect.isChecked()):
                    book.etype = Book.TYPE_DIRECT
                else:
                    book.etype = Book.TYPE_DONATION
                book.writetosheet(Book.DATA_FILE)
                
                self.parent.loadbooks()

                self.reset(book.id)
            except:
                print("That's an error")
    def reset(self,id):
        self.dataTitle.setText("")
        self.dataID.setText(str(id+1))
        self.dataAuthor.setText("")
        self.dataSubject.setText("")
        self.dataLanguage.setText("")
        self.dataPublisher.setText("")
        self.dataYear.setText("")
        self.dataPrice.setText("")

class BookEdit(QDialog):
    def __init__(self,parent,book):
        super(BookEdit, self).__init__()
        uic.loadUi('res/bookentry.ui', self)
        self.setFixedSize(self.size())
        self.parent = parent
        self.book = book
        self.submit.clicked.connect(self.onupdate)
        try:
            self.dataTitle.setText(book.title)
            self.dataID.setText(str(book.id))
            self.dataAuthor.setText(book.author)
            self.dataSubject.setText(book.subject)
            self.dataLanguage.setText(book.language)
            self.dataPrice.setText(str(book.price))
            self.dataPublisher.setText(book.publisher)
            self.dataYear.setText(str(book.year))
            self.radioReference.setChecked(book.booktype ==  Book.BOOKTYPE_REFERENCE)
            self.radioGeneral.setChecked(book.booktype ==  Book.BOOKTYPE_GENERAL)
            self.radioGift.setChecked(book.etype == Book.TYPE_GIFT)
            self.radioDirect.setChecked(book.etype == Book.TYPE_DIRECT)
            self.radioDonation.setChecked(book.etype == Book.TYPE_DONATION)
        except:
            print("That's an error")
        self.show()
    def onupdate(self):
        if self.dataTitle.text() != "" and self.dataID.text() != "":
            #try:
                book = Book()
                book.date = [datetime.now().day,datetime.now().month,datetime.now().year]
                book.title = self.dataTitle.text()
                book.id = eval(self.dataID.text())
                book.author = self.dataAuthor.text()
                book.subject = self.dataSubject.text()
                book.language = self.dataLanguage.text()
                book.price = eval(self.dataPrice.text())
                book.publisher = self.dataPublisher.text()
                book.year = eval(self.dataYear.text())
                if(self.radioReference.isChecked()):
                    book.booktype = Book.BOOKTYPE_REFERENCE
                else:
                    book.booktype = Book.BOOKTYPE_GENERAL
                if(self.radioGift.isChecked()):
                    book.etype = Book.TYPE_GIFT
                elif(self.radioDirect.isChecked()):
                    book.etype = Book.TYPE_DIRECT
                else:
                    book.etype = Book.TYPE_DONATION
                book.replace(Book.DATA_FILE)
                self.parent.loadbooks()
                self.hide()
            #except:
            #    print("error")
        pass

        
class Book:
    BOOKTYPE_GENERAL = 10
    BOOKTYPE_REFERENCE = 11
    DATA_FILE = "data/books.xlsx"
    TYPE_DIRECT = 0
    TYPE_DONATION = 1
    TYPE_GIFT = 2

    def __init__(self):
        self.date = [0,0,0]
        self.title = ""
        self.id = 0
        self.subject = ""
        self.language = ""
        self.price = 0
        self.publisher = ""
        self.year = 0
        self.issued = 0
        self.booktype = 0
        self.etype = 0
        self.author = ""
        self.issued = "none"
        self.row = 0
    def loadfromrow(row):
        rt = Book()
        dt=str(row[0].value).split("-")
        rt.date[0]=eval(dt[0])
        rt.date[1]=eval(dt[1])
        rt.date[2]=eval(dt[2])
        rt.title =str(row[1].value)
        rt.id = eval(str(row[2].value))
        rt.author = str(row[3].value)
        rt.subject = str(row[4].value)
        rt.language = str(row[5].value)
        rt.price = eval(str(row[6].value))
        rt.publisher = (str(row[7].value))
        rt.year = eval(str(row[8].value))
        rt.booktype = eval(str(row[9].value))
        rt.etype = eval(str(row[10].value))
        rt.issued = str(row[11].value)
 
        return rt
    def writetosheet(self,path):
        sheet = load_workbook(path)
        last = len(sheet["Main"]["A"])
        for i in range(1,last):
            if(sheet["Main"][i][0].value == None  or sheet["Main"][i][0].value == ""):
                last = i-1
                break
        row = sheet["Main"][last+1]
        row[0].value = str(self.date[0])+"-"+str(self.date[1])+"-"+str(self.date[2])
        row[1].value = self.title
        row[2].value = self.id
        row[4].value = self.subject
        row[5].value = self.language
        row[6].value = self.price
        row[7].value = self.publisher
        row[8].value = self.year
        row[9].value = self.booktype
        row[10].value = self.etype
        row[3].value = self.author
        row[11].value = self.issued
        sheet.save(path)
    def delete(self,path):
        book = load_workbook(path)
        sheet = book["Main"]
        if(self.row == 0):
            for i in range(1,len(sheet["A"])):
                try:
                    bk = Book.loadfromrow(sheet[i])
                    if(bk.id == self.id): self.row = i
                except:
                    pass
        sheet.delete_rows(self.row)
        book.save(path)
    def replace(self,path):
        book = load_workbook(path)
        sheet = book["Main"]
        if(self.row == 0):
            for i in range(1,len(sheet["A"])):
                try:
                    bk = Book.loadfromrow(sheet[i])
                    if(bk.id == self.id): self.row = i
                except:
                    pass
        row = sheet[self.row]
        row[0].value = str(self.date[0])+"-"+str(self.date[1])+"-"+str(self.date[2])
        row[1].value = self.title
        row[2].value = self.id
        row[4].value = self.subject
        row[5].value = self.language
        row[6].value = self.price
        row[7].value = self.publisher
        row[8].value = self.year
        row[9].value = self.booktype
        row[10].value = self.etype
        row[3].value = self.author
        row[11].value = self.issued
        book.save(path)
        
