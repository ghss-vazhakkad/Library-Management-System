import sys
import typing
from PyQt5 import QtCore, uic
from PyQt5.QtWidgets import *
from PyQt5.QtWidgets import QWidget
from openpyxl import Workbook, load_workbook

class AddMember(QDialog):
    def __init__(self,parent):
        super(AddMember,self).__init__()
        uic.loadUi('res/addmember.ui', self)
        self.setFixedSize(self.size())
        self.show()
        self.parent = parent
        self.addButton.clicked.connect(self.addmember)
        self.reset()
    def addmember(self):
        name = str(self.name.text())
        id = 0
        try:
            id = eval(self.id.text())
        except:
            pass
        member = Member()
        if(name != ""):
            member.name = name
            member.id = id
            if(self.radioTeacher.isChecked()):
                member.status = Member.STATUS_TEACHER
            elif(self.radioNonTeacher.isChecked()):
                member.status = Member.STATUS_NON_TEACHER
            elif(self.radioPlusOne.isChecked()):
                member.status = Member.STATUS_PLUS_ONE
            else:
                member.status = Member.STATUS_PLUS_TWO
            member.save("data/members.xlsx")
            self.parent.loadmembers()
            self.reset()
    def reset(self):
        self.name.setText("")
        self.id.setText(str(self.parent.getLastMemberID()))


class EditMember(QDialog):
    def __init__(self,parent,member):
        super(EditMember,self).__init__()
        uic.loadUi('res/addmember.ui', self)
        self.setFixedSize(self.size())
        self.show()
        self.parent = parent
        self.addButton.setText("UPDATE")
        self.addButton.clicked.connect(self.addmember)

        self.name.setText(member.name)
        self.id.setText(str(member.id))
        radios = [self.radioTeacher,self.radioNonTeacher,self.radioPlusOne,self.radioPlusTwo]
        statuses = [Member.STATUS_TEACHER,Member.STATUS_NON_TEACHER,Member.STATUS_PLUS_ONE,Member.STATUS_PLUS_TWO]
        i=0
        while i < 4:
            radios[i].setChecked(False)
            if(member.status == statuses[i]): radios[i].setChecked(True)
            i += 1
        self.last = member
    def addmember(self):
        name = str(self.name.text())
        id = 0
        try:
            id = eval(self.id.text())
        except:
            pass
        member = self.last
        if(name != ""):
            member.name = name
            member.id = id
            if(self.radioTeacher.isChecked()):
                member.status = Member.STATUS_TEACHER
            elif(self.radioNonTeacher.isChecked()):
                member.status = Member.STATUS_NON_TEACHER
            elif(self.radioPlusOne.isChecked()):
                member.status = Member.STATUS_PLUS_ONE
            else:
                member.status = Member.STATUS_PLUS_TWO
            member.update("data/members.xlsx")
            self.parent.loadmembers()
            self.hide()


class Member:
    STATUS_TEACHER = "Teacher"
    STATUS_NON_TEACHER = "Non-Teaching Staff"
    STATUS_PLUS_TWO = "Plus Two"
    STATUS_PLUS_ONE = "Plus One"
    def __init__(self):
        self.name = ""
        self.id = 0
        self.status = Member.STATUS_PLUS_TWO
        self.row = 0
    def save(self,path):
        sheet = load_workbook(path)
        last = len(sheet["Sheet1"]["A"])
        for i in range(1,last):
            if(sheet["Sheet1"][i][0].value == None  or sheet["Sheet1"][i][0].value == ""):
                last = i-1
                break
        row = sheet["Sheet1"][last+1]
        if(self.id == 0): self.id = last+1
        row[0].value = self.id
        row[1].value = self.name
        row[2].value = self.status
        sheet.save(path)
    def load(path):
        rt = []
        sheet = load_workbook(path)["Sheet1"]
        length = len(sheet["A"])
        for i in range(1,length):
            sample = str(sheet[i+1][0].value)
            if(sample != "None" and sample != ""):
                mem = Member()
                row = sheet[i+1]
                mem.row = i+1
                mem.id = row[0].value
                mem.name = str(row[1].value)
                mem.status = str(row[2].value)
                rt.append(mem)
        return rt
    def update(self,path):
        sheet = load_workbook(path)
        sh = sheet["Sheet1"]
        length = len(sh["A"])
        row = sh[self.row]
        row[0].value = self.id
        row[1].value = self.name
        row[2].value = self.status

        sheet.save(path)
        
    def delete(self,path):
        book = load_workbook(path)
        sheet = book["Sheet1"]
        sheet.delete_rows(self.row)
        book.save(path)
                
    def getMemberById(id,path):
        members = Member.load(path)
        for member in members:
            if(member.id == id):
                return member
        return None
    def fgetMemberById(id,members):
        for member in members:
            if(member.id == id):
                return member
        return None
